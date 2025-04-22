# %%
import matplotlib.patches as patches
import scipy.stats as stats
import numpy as np
import matplotlib.pyplot as plt
from my_tools.pyplot_helper import *
import joblib
from pathlib import Path
from my_tools import data_location
from my_tools.color_preset import *
from openpyxl import Workbook
from datetime import date
#
# %%


def significant_coactivated_pairs(data_root, rat, target_pairs):
    _, significance, region, pair_id, * \
        _ = joblib.load(
            data_root / rat / f"analyses/{rat}-coact_sig-cue_and_extinction-ext.joblib")

    cnt = np.zeros((target_pairs.shape[0], 2, 5))
    sig = []
    for n, target_pair in enumerate(target_pairs):
        idx = np.where(np.all(region.T == target_pair, axis=1) |
                       np.all(region.T == target_pair[::-1], axis=1))[0]
        sig.append(significance[:, idx] == 1)
        cnt[n, 0, :] = (significance[:, idx] == 1).sum(axis=1)
        cnt[n, 1, :] = idx.shape[0]

    return cnt, sig


# %%
def get_data_3bc():
    # parameters
    data_root = data_location.get_rootdir()
    rats = data_location.get_rats()
    target_hc = [2, 3]

    count = np.zeros((len(rats), 3, 2, 5))
    significance = [np.zeros([5, 0]), np.zeros([5, 0]), np.zeros([5, 0])]
    target_pairs = np.array(
        [["BLA", "PL5"], ["vCA1", "PL5"], ["BLA", "vCA1"],])
    for rat_index, rat in enumerate(rats):
        count[rat_index], temp = significant_coactivated_pairs(
            data_root, rat, target_pairs)
        for n in range(len(temp)):
            significance[n] = np.hstack((significance[n], temp[n]))

    for n in range(3):
        significance[n] = significance[n][target_hc, :].astype(bool)
    count = count[:, :, :, target_hc].sum(axis=0)

    p_values = np.zeros(count.shape[0])
    odds_ratio = np.zeros(count.shape[0])
    n_sample = np.zeros((count.shape[0], count.shape[2]))
    for n in range(count.shape[0]):
        n_sample[n,:]=count[n, 1, :]
        subset=count[n, :, :]
        subset[1, :] = subset[1, :] - subset[0, :]
        odds_ratio[n], p_values[n], *_ = stats.fisher_exact(subset)

    # c = np.zeros((len(target_pairs), 2000))
    actual = np.zeros(len(target_pairs))
    for idx in range(len(target_pairs)):
        n = int(count[idx, 1, 0])
        p1 = count[idx, 0, 0]/count[idx, 1, 0]
        p2 = count[idx, 0, 1]/count[idx, 1, 1]
        # for ite in range(2000):
        #     c[idx, ite] = (np.random.choice([True, False], size=n, p=[
        #         p1, 1-p1]) & np.random.choice([True, False], size=n, p=[p2, 1-p2])).sum()
        actual[idx] = (significance[idx][0, :] & significance[idx][1, :]).sum()

    overlap = np.array([[
        (significance[idx][0, :] & significance[idx][1, :]).sum(),
        (significance[idx][0, :] & ~significance[idx][1, :]).sum(),
        (~significance[idx][0, :] & significance[idx][1, :]).sum(),
        (~significance[idx][0, :] & ~significance[idx][1, :]).sum(),
    ]for idx in range(len(target_pairs))])

    data_b = target_pairs, count, p_values, odds_ratio, n_sample
    data_c = target_pairs, overlap
    return data_b, data_c

# %%


def get_sig_mark(p_value):
    if p_value < 0.001:
        return "***"
    elif p_value < 0.01:
        return "**"
    elif p_value < 0.05:
        return "*"
    else:
        return "N.S."


# %%
def plot_3b(fig, x, y, data, fontsize=8,
            colors={
                frozenset(('BLA', 'PL5')): 'blue',
                frozenset(('vCA1', 'PL5')): 'orange',
                frozenset(('BLA', 'vCA1')): 'green'
            }):
    width = 60
    height = 25
    thershold = 0.5

    target_pairs, count, p_values, odds_ratio, n_sample = data
    line_width = 0.5
    ax = subplot_mm([x, y, width, height])
    for idx, target_pair in enumerate(target_pairs):
        basic_color = colors[frozenset(target_pair)]
        light_color = lighten_color(basic_color, brightness=0.7)
        basic_color = lighten_color(basic_color, brightness=0.4)
        y_val = np.array(
            [count[idx, 0, n]/count[idx, 1, n]*100 for n in range(2)])

        col = [light_color, basic_color]
        for n in range(2):
            ax.bar(idx*2+0.6+0.8*n, y_val[n],
                   width=0.7, color=col[n], edgecolor=basic_color,
                   linewidth=line_width, zorder=2)
            ax.text(idx*2+0.6+0.8*n,
                    y_val[n]-0.2, int(count[idx, 0, n]), ha="center", va="top")

        # ax.bar(idx*2+1.4, count[idx, 0, 1]/count[idx, 1, 1]*100,
        #        width=0.7, color=basic_color,
        #        edgecolor=basic_color, linewidth=line_width, zorder=2)
        # ax.text(idx*2+1.4, count[idx, 0, 1]/count[idx, 1, 1] *
        #         100-0.2, int(count[idx, 0, 1]), ha="center", va="top")

        ax.plot(idx*2+np.array([0.6, 0.6, 1.4, 1.4]),
                0.2+np.array([
                    y_val[0],
                    np.max(y_val)+0.2,
                    np.max(y_val)+0.2,
                    y_val[1]]),
                'k-', linewidth=line_width)
        if p_values[idx] < 0.05:
            font_magnification = 1.5
        else:
            font_magnification = 1.0

        ax.text(idx*2+1,
                np.max(y_val)+0.5,
                get_sig_mark(p_values[idx]), ha="center",
                va="bottom", size=fontsize*font_magnification)
    ax.set_ylim(0, 8)
    x_range = ax.get_xlim()
    ax.set_xlim(x_range)
    ax.plot(x_range, thershold*np.ones(2), '-', zorder=1,
            lw=0.5, color=0.5*np.ones(3))
    box_off()
    ax.set_ylabel("Fraction of\ncoactivated\nensemble pairs (%)")
    ax.set_xticks(np.arange(len(target_pairs))*2+1)
    ax.set_xticklabels([x[0] + ' - ' + x[1] for x in target_pairs])

    label_strs = ["Pre-extinction sleep", "Post-extinction sleep"]
    top_positions = [7, 6.1]
    face_colors = [0.7, 0.4]
    edge_colors = [0.4, 0.4]
    height = 0.6
    width = 0.5
    left_position = 3
    for top_posision, label_str, face_color, edge_color in\
            zip(top_positions, label_strs, face_colors, edge_colors):
        rc = patches.Rectangle((left_position, top_posision), width, height, facecolor=face_color*np.ones(3),
                               edgecolor=edge_color*np.ones(3), linewidth=line_width)
        ax.add_patch(rc)
        ax.text(left_position+width*1.1, top_posision+height /
                2, label_str, ha="left", va="center")

# %%


def stats_3b(ws, data):
    target_pairs, count, p_values, odds_ratio, n_sample = data
    ws.append(["Region pair", "Statistical test","Statistical value","P-value"])
    excel_fonts = excel_font_preset()
    for cell in ws[1]:
        cell.font = excel_fonts["heading"]

    for idx, target_pair in enumerate(target_pairs):
        n_str = ", ".join([f"{int(x)}" for x in n_sample[idx]])
        ws.append(
            [f"{target_pair[0]} - {target_pair[1]}",
            f"Fisher's exact test (n = {n_str})",
            f"Odds ratio = {odds_ratio[idx]:.3f}",
            f"{p_values[idx]:.3f}"])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = excel_fonts["default"]
    excel_auto_expand(ws)
# %%


def plot_3c(fig, x, y, data, fontsize=8,
            colors={
                frozenset(('BLA', 'PL5')): 'blue',
                frozenset(('vCA1', 'PL5')): 'orange',
                frozenset(('BLA', 'vCA1')): 'green'
            }):
    target_pairs, overlap = data
    width = 22
    height = width
    hatches = ['//', None]
    line_width = 0.5
    for idx, target_pair in enumerate(target_pairs):
        ax = subplot_mm([x+width*idx, y, width, height])
        basic_color = colors[frozenset(target_pair)]
        light_color = lighten_color(basic_color, brightness=0.7)
        basic_color = lighten_color(basic_color, brightness=0.4)
        wedges, *_ = ax.pie(
            overlap[idx, [0, 2]],
            labels=[int(x) for x in overlap[idx, [0, 2]]],
            labeldistance=0.6,
            colors=[basic_color],
            startangle=90,
            wedgeprops=dict(
                edgecolor=(light_color),
                linewidth=line_width))
        for wedge, hatch in zip(wedges, hatches):
            wedge.set_hatch(hatch)

        ax.set_title(f"{target_pairs[idx,0]} - {target_pairs[idx,1]}",
                     fontsize=fontsize,
                     y=-0.15)

        if idx == 0:
            ax.set_ylabel(
                "Among ensemble pairs\ncoactivated in \npost-extinction sleep")

    ax = subplot_mm([x, y+height+2, width*3, 6])
    ax.set_xlim(0, width*3)
    ax.set_ylim(0, 6)

    pair_types = ["Unchanged", "Reorganized"]
    hatches = ["//", None]
    left_positions = [7, 33]
    for left_posisiton, pair_type, hatch in zip(left_positions, pair_types, hatches):
        rc = patches.Rectangle((left_posisiton, 2), 8, 3, facecolor=0.4*np.ones(3),
                               edgecolor=0.7*np.ones(3), linewidth=line_width)
        rc.set_hatch(hatch)
        ax.add_patch(rc)
        ax.text(left_posisiton+9, 3.5, pair_type, ha="left", va="center")
    ax.set_axis_off()
# %%


def get_data_3a():
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()

    # rat = rats[8]
    # pair_idx=121
    rat = rats[7]
    pair_idx = 434

    template = "CueAndExtinction-Ext"

    _, basic_info, session_info, *_ \
        = joblib.load(data_dir / rat / (rat + "-basic_info.joblib"))

    _, ica_weight, *_ = joblib.load(data_dir / rat /
                                    'analyses' / (rat + '-ica_weight.joblib'))

    _, spikes, cluster_info, *_ \
        = joblib.load(data_dir / rat / (rat + "-ok_spikes.joblib"))

    _, react_strength, react_param = \
        joblib.load(data_dir / rat / 'analyses' /
                    (rat + "-ica_react_strength-cue_and_extinction-ext.joblib"))

    react_time = (
        np.arange(react_strength.shape[0])+0.5)*react_param["tbin_size"]

    _, ccg, shuffle, params = joblib.load(
        data_dir / rat / "analyses" / f"{rat}_icaReacZNCCG_{template}.joblib")

    template_idx = 5
    hc_idx = 3

    react_ids = params["pair_id"][:, pair_idx]-1
    weights = ica_weight[template_idx].iloc[react_ids]["weight"].values
    regions = ica_weight[template_idx].iloc[react_ids]["region"].values
    cell_id = ica_weight[template_idx].iloc[react_ids]["cell_id"].values
    cell_id = [np.array(x) for x in cell_id]

    target_ccg = ccg["ccg"][hc_idx, :, pair_idx]
    ci99 = shuffle["global99"][:, :, pair_idx]
    ci99 = ci99[hc_idx, :]
    ci99 = np.reshape(ci99, (1, 2))

    n_bin_half = (ccg["ccg"].shape[1]-1)/2
    ccg_t = np.arange(-n_bin_half, n_bin_half+1)*params["tBinSize"]*1000

    target_ccg = target_ccg[abs(ccg_t) <= 500]
    ccg_t = ccg_t[abs(ccg_t) <= 500]

    sorted_weights = []
    sorted_cell_id = []
    id_map = {}
    min_cell = 0
    for i in range(len(weights)):
        sorted_indices = np.argsort(weights[i])[::-1]
        sorted_weights.append(weights[i][sorted_indices])
        sorted_cell_id.append(cell_id[i][sorted_indices])

        for new_id, old_id in enumerate(sorted_cell_id[i]):
            id_map[old_id] = new_id + min_cell

        sorted_cell_id[i] = np.array([id_map[n] for n in sorted_cell_id[i]])
        min_cell += len(weights[i])

    z_scored_weight = [(x - x.mean())/x.std() for x in sorted_weights]

    inst_strength = react_strength[:, react_ids]

    return \
        inst_strength, react_time, \
        spikes, id_map, sorted_cell_id, z_scored_weight, sorted_weights, regions, \
        target_ccg, ccg_t, ci99, react_ids

# %%


def plot_3a(fig, x, y, data, fontsize=8,
            color={"BLA": "#CC0000", "vCA1": "#0000CC", "PL5": "#00CC00"},
            ccg_color={
        frozenset(("BLA", "PL5")): "#CCCC00",
        frozenset(("vCA1", "PL5")): "#00CCCC",
        frozenset(("BLA", "vCA1")): "#CC00CC"}):
    raster_width = 55
    react_height = 8
    raster_height = 25
    weight_width = 10
    x_gap = 3
    ccg_gap_x = 14
    ccg_gap_y = 3
    ccg_width = 25
    ccg_height = 20
    duration = 2

    t_starts = [43262.04, 42998.88]

    inst_strength, react_time, \
        spikes, id_map, sorted_cell_id, z_scored_weight, sorted_weights, regions, \
        target_ccg, ccg_t, ci99, react_ids = data

    cell_to_show = id_map.keys()
    sig_cell_id = []
    nsig_cell_id = []
    threshold = 2.0
    for n in range(len(z_scored_weight)):
        sig_cell_id.append(sorted_cell_id[n][z_scored_weight[n] > threshold])
        nsig_cell_id.append(sorted_cell_id[n][z_scored_weight[n] <= threshold])

    col_map = {}
    for i in range(len(z_scored_weight)):
        for cell_id, weight in zip(sorted_cell_id[i], z_scored_weight[i]):
            col_map[cell_id] = color[regions[i]] if weight > 2 else lighten_color(
                color[regions[i]], 0.8)

    ones = np.ones((2, 1))
    for n, t_start in enumerate(t_starts):
        t_end = t_start+duration
        react_time_idx = (react_time >= t_start) & (react_time < t_end)

        idx = (spikes["spiketime"] >= t_start) & (spikes["spiketime"] <= t_end)
        spk = spikes["spiketime"][idx].values
        clu = spikes["cluster"][idx].values

        idx = np.where([x in cell_to_show for x in clu])[0]
        spk = spk[idx]
        clu = np.array([id_map[c] for c in clu[idx]])

        ax = subplot_mm([x + (raster_width+x_gap)*n, y,
                        raster_width, react_height], fig=fig)
        for i in range(2):
            ax.plot(react_time[react_time_idx]-t_start, inst_strength[react_time_idx, i]-i*5,
                    c=color[regions[i]], linewidth=0.75)
        ax.set_xlim(0, t_end-t_start)
        ax.set_ylim(-10, 90)
        ax.plot([duration*0.75, duration*0.85], 20*ones, 'k-', lw=1.5)
        ax.text(
            duration*0.8, 21, f"{duration*0.1*1000:.0f} ms",
            fontsize=fontsize, ha="center", va="bottom")
        ax.plot(duration*0.69*ones, [21, 51], 'k-', lw=1.5)
        ax.text(
            duration*0.68, 36, "30 z", fontsize=fontsize, ha="right", va="center")
        ax.set_axis_off()

        ax = subplot_mm([x+(raster_width+x_gap)*n, y+react_height,
                        raster_width, raster_height], fig=fig)
        col = [col_map[c] for c in clu]
        ax.scatter(spk-t_start, clu, s=0.25, c=col)

        ax.invert_yaxis()
        ax.set_xlim(0, t_end-t_start)
        ax.set_axis_off()

        if n == 0:
            for i in range(len(regions)):
                ax.text(-duration*0.04, sorted_cell_id[i].mean(),
                        regions[i],
                        color=color[regions[i]],
                        rotation=90,
                        fontsize=fontsize, ha="center", va="bottom")

    ax = subplot_mm([x+2*(raster_width+x_gap)+ccg_gap_x, y +
                    ccg_gap_y, ccg_width, ccg_height], fig=fig)

    ax.plot(ccg_t, target_ccg, color=lighten_color(ccg_color[frozenset(regions)], 0.4),
            lw=0.75, zorder=2)
    ax.plot(ccg_t[[0, -1]], ones@ci99, '-',
            color=lighten_color(ccg_color[frozenset(regions)], 0.7), lw=0.75, zorder=1)
    ax.set_xlim(-500, 500)
    ylim = ax.get_ylim()
    ax.set_ylim(ylim)
    ax.set_yticks(np.arange(0, ylim[1], 0.03))

    ax.plot([0, 0], ylim, '-', color="#999999", lw=0.75, zorder=1)
    box_off()
    ax.set_xlabel("$\Delta$ time (ms)", fontsize=fontsize)
    ax.set_ylabel("Correlation (r)", fontsize=fontsize)
# %%

data_3a = get_data_3a()
data_3b, data_3c = get_data_3bc()
# %%
fontsize = 7
panel_fontsize = 9
plt.rcParams['font.size'] = fontsize
default_hatch_linewidth = plt.rcParams['hatch.linewidth']
plt.rcParams['hatch.linewidth'] = 3.5

colors = color_preset("region_pairs")

fig_3 = fig_mm([174, 71])
# draw_fig_border()

raster_color = color_preset("regions")
ccg_color = color_preset("region_pairs")
x = 11
y = 2
plot_3a(fig_3, x, y, data_3a, fontsize=fontsize,
        ccg_color=ccg_color, color=raster_color)
text_mm((x-8, y+1), 'A', fontsize=panel_fontsize, weight='bold', fig=fig_3)

x = 20
y = 40
plot_3b(fig_3, x, y, data_3b, fontsize=fontsize, colors=colors)
text_mm((x-17, y), 'B', fontsize=panel_fontsize, weight='bold', fig=fig_3)

x = 100
y = 39
plot_3c(fig_3, x, y, data_3c, fontsize=fontsize, colors=colors)
text_mm((x-13, y+1), 'C', fontsize=panel_fontsize, weight='bold', fig=fig_3)

plt.show()
fig_dir = Path('~/Dropbox/coact_stability_paper/figures').expanduser()
stats_dir = Path('~/Dropbox/coact_stability_paper/stats').expanduser()

if not fig_dir.exists():
    fig_dir.mkdir(parents=True)
fig_3.savefig(fig_dir / f'fig_3_{date.today().strftime("%Y_%m%d")}.pdf')

plt.rcParams['hatch.linewidth'] = default_hatch_linewidth

wb = Workbook()
ws = wb.active
ws.title = "Figure 3B"
stats_3b(ws, data_3b)
if not stats_dir.exists():
    stats_dir.mkdir(parents=True)
wb.save(stats_dir / f'fig_3_{date.today().strftime("%Y_%m%d")}.xlsx')


# %%

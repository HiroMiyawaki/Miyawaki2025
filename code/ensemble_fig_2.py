# %%
import joblib
import numpy as np
from pathlib import Path
from my_tools import data_location
from my_tools.pyplot_helper import *
from my_tools.color_preset import *
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from datetime import datetime as dt
import pandas as pd
import scipy
from scipy import stats
import scikit_posthocs as sp
import itertools
from matplotlib import font_manager
from rpy2.robjects import pandas2ri
import rpy2.robjects as ro
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
# %%


def get_data_2a():
    # load data
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()

    param = {}
    for idx, rat in enumerate(rats):
        _, temp, param[rat] = joblib.load(
            data_dir / rat / "analyses" / (rat + "-ica_type-cue_and_extinction-ext.joblib"))
        temp.insert(0, 'rat', rat)
        if idx == 0:
            ic_type = temp.copy()
        else:
            ic_type = pd.concat((ic_type, temp))

    ic_type = ic_type.reset_index(drop=True)
    ic_type['link'] = ic_type['link'].str.capitalize()

    # summarize data
    link_typename = ic_type.link.unique()

    type_count = pd.DataFrame([], index=link_typename)
    regions = ['BLA', 'PL5', 'vCA1']
    for region in regions:
        temp = ic_type.query('region == @region').value_counts('link')
        type_count[region] = temp

    p_values = {}
    degrees_of_freedom = {}
    chi_squared = {}
    for r_idx in itertools.combinations(range(len(regions)), 2):
        chi2, p, df, \
            _ = scipy.stats.chi2_contingency(
                type_count[[regions[x] for x in r_idx]])
        p_values[frozenset(regions[x] for x in r_idx)] = p * \
            scipy.special.comb(len(regions), 2)
        chi_squared[frozenset(regions[x] for x in r_idx)] = chi2
        degrees_of_freedom[frozenset(regions[x] for x in r_idx)] = df

    return type_count, regions, p_values, chi_squared, degrees_of_freedom


def plot_2a(fig, x, y, data, fontsize=8,
            colors={"maintained": "blue",
                    "initiated": "orange",
                    "terminated": "green",
                    "transient": "red"},
            plot_order=["Maintained", "Initiated", "Terminated", "Transient"]):
    type_count, regions, p_values, chi_squared, degrees_of_freedom = data

    colors = [colors[x.lower()] for x in plot_order]
    pie_width = 18
    pie_height = pie_width
    bar_height = 3

    for r_idx, region in enumerate(regions):
        ax = subplot_mm([x+pie_width*r_idx, y, pie_width, pie_height], fig=fig)
        target_pd = type_count[region].reindex(plot_order)

        target_pd.plot.pie(
            startangle=90, ylabel='',
            labeldistance=0.7,
            labels=target_pd.values,
            counterclock=False,
            fontsize=fontsize, colors=colors)
        for text in ax.texts:
            if hasattr(text, 'get_text') and '%' not in text.get_text():
                text.set_ha('center')
                text.set_va('center')
        ax.set_title(region, size=fontsize, y=0.9)

    ax = subplot_mm([x, y+pie_height-1, pie_width*3, bar_height], fig=fig)
    ax.set_xlim(0, 90)
    ax.set_ylim(0, 5)

    for r_idx in itertools.combinations(range(len(regions)), 2):
        key = frozenset(regions[x] for x in r_idx)
        p = p_values[key]

        if p < 0.05:
            ax.plot(np.array([r_idx[x] for x in [0, 0, 1, 1]])*30+15+[0.5,
                    0.5, -0.5, -0.5], np.array([1, 0, 0, 1])*2+2, 'k-', lw=0.75)
            if p < 0.001:
                sig_txt = '***'
            elif p < 0.01:
                sig_txt = '**'
            elif p < 0.05:
                sig_txt = '*'
            else:
                sig_txt = 'NS'
            ax.text(np.mean(r_idx)*30+15, 2, sig_txt,
                    ha='center', va='top', size=fontsize*1.5)
    ax.axis('off')

# %%


def plot_2a_legend(fig, x, y, fontsize=8, colors={"maintained": "blue",
                                                  "initiated": "orange",
                                                  "terminated": "green",
                                                  "transient": "red"},
                   plot_order=["Maintained", "Initiated", "Terminated", "Transient"]):

    cell_height = (fontsize)*0.3528*1.15
    row_label_width = 0.3528*fontsize*8*1.15
    cell_width = 0.3528*fontsize*5*1.15
    column_label_height = (fontsize+1)*0.3528

    font_props = font_manager.FontProperties(family="Sans")

    status = {
        "maintained": [True, True],
        "initiated": [False, True],
        "terminated": [True, False],
        "transient": [False, False]}
    # status_str = ["\u2716No", "\u2714Yes"]
    status_str = ["No", "Yes"]
    str_color = ["#cc6677", "#33aa88"]

    column_labels = ["Conditioning", "Retention"]

    ax = subplot_mm([x, y, row_label_width+cell_width*2,
                    cell_height*len(plot_order)+2*column_label_height], fig=fig)
    ax.set_xlim(0, row_label_width+cell_width*2)
    ax.set_ylim(0, cell_height*len(plot_order)+2*column_label_height)

    ax.add_patch(
        patches.Rectangle(
            (row_label_width, 0),
            cell_width*2, column_label_height,
            facecolor='white', edgecolor='none'))

    ax.text(row_label_width+cell_width, column_label_height/2,
            "Detected in", va='center', ha='center', fontsize=fontsize)

    ax.text(row_label_width/2, column_label_height,
            "Extinction-ensemble\ncategory", va='center', ha='center', fontsize=fontsize)

    for n, label in enumerate(column_labels):
        ax.add_patch(
            patches.Rectangle(
                (row_label_width+n*cell_width, column_label_height),
                cell_width, column_label_height,
                facecolor='white', edgecolor='none'))
        ax.text(
            row_label_width+n*cell_width+cell_width/2,
            column_label_height*1.5,
            label,
            va='center', ha='center',
            fontsize=fontsize)

    for idx, label in enumerate(plot_order):
        ax.add_patch(
            patches.Rectangle(
                (0, idx*cell_height+2*column_label_height),
                row_label_width, cell_height,
                facecolor=colors[label.lower()], edgecolor='none'))
        ax.text(
            0.5*row_label_width,
            (idx+0.5)*cell_height + 2*column_label_height,
            label,
            va='center', ha='center',
            fontsize=fontsize)
        for n, state in enumerate(status[label.lower()]):
            ax.add_patch(
                patches.Rectangle(
                    (row_label_width+cell_width*n, idx *
                     cell_height+2*column_label_height),
                    cell_width, cell_height,
                    facecolor=lighten_color(colors[label.lower()]), edgecolor='none'))
            ax.text(
                row_label_width+cell_width*(n+0.5),
                (idx+0.5)*cell_height+2*column_label_height,
                status_str[state],
                color=str_color[state],
                va='center', ha='center',
                fontproperties=font_props,
                fontsize=fontsize)
    ax.invert_yaxis()
    unity = np.ones(2)
    ax.plot(row_label_width*unity, [0, cell_height*len(plot_order)+2*column_label_height], '-',
            lw=0.5, color="#000000")
    ax.plot((row_label_width+cell_width)*unity,
            [column_label_height, cell_height *
                len(plot_order)+2*column_label_height], '-',
            lw=0.5, color="#000000")
    ax.plot([0, row_label_width+cell_width*2],
            2*column_label_height*unity, '-',
            lw=0.5, color="#000000")
    # ax.axis('off')
    ax.set_xticks([])
    ax.set_yticks([])

# %%


def stats_2a(ws, data):
    type_count, regions, p_values, chi_squared, degrees_of_freedom = data
    excel_fonts = excel_font_preset()

    ws.append(["Region pair", "Statistical test", "Statistical value",
               "P-value without Bonferroni correction"])

    for cell in ws[1]:  # 1行目のセルすべてにアクセス
        cell.font = excel_fonts["heading"]  # ヘッダーのフォントを変更

    for r_idx in itertools.combinations(range(len(regions)), 2):
        key = frozenset(regions[x] for x in r_idx)
        p = p_values[key]

        if p < 0.001:
            p_str = f"{p:.3e}"
        else:
            p_str = f"{p:.3f}"
        n_ensembles = ', '.join(
            [str(type_count[regions[x]].sum()) for x in r_idx])
        ws.append([
            " - ".join([regions[x] for x in r_idx]),
            f"Chi-squared test (n = {n_ensembles})",
            f"χ²₍{to_subscript(degrees_of_freedom[key])}₎ = {chi_squared[key]:.2f}",
            p_str])

    for row in ws.iter_rows(min_row=2):  # 2行目以降を対象
        for cell in row:
            cell.font = excel_fonts["default"]

    excel_auto_expand(ws)


# %%

def get_data_2b():
    # lod data
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()

    param = {}
    for idx, rat in enumerate(rats):
        _, temp, param[rat] = joblib.load(
            data_dir / rat / "analyses" / (rat + "-ica_type-cue_and_extinction-ext.joblib"))
        temp.insert(0, 'rat', rat)
        if idx == 0:
            ic_type = temp.copy()
        else:
            ic_type = pd.concat((ic_type, temp))

    ic_type = ic_type.reset_index(drop=True)

    # summarize data
    # link_typename = ic_type.link.unique()
    link_typename = ['maintained', 'initiated', 'terminated', 'transient']
    type_count = pd.DataFrame([], index=link_typename)
    regions = ['BLA', 'PL5', 'vCA1']
    for region in regions:
        temp = ic_type.query('region == @region').value_counts('link')
        type_count[region] = temp

    activation_rate = {}
    p_friedman = {}
    p_posthock = {}
    stats_posthoc = {}
    chi2_friedman = {}
    ro.r('library("PMCMRplus")')
    pandas2ri.activate()
    for region in regions:
        for link in link_typename:
            target_pd = ic_type.query(
                'region == @region and link == @link')[[f'hc{x}_nrem' for x in range(1, 5)]]
            target = target_pd.to_numpy() * 60
            activation_rate[region, link] = target

            res = stats.friedmanchisquare(*target.T)
            chi2_friedman[region, link] = res[0]

            p_friedman[region, link] = res[1]
            target_pd = target_pd.reset_index()  # indexを列にする（列名はデフォルトで 'index'）
            target_pd = target_pd.melt(
                id_vars='index',         # 被験者IDの列を固定
                var_name='hc',        # 新しい"条件(群)名"列
                value_name='value'       # 実測値が入る列名
            )
            target_pd = target_pd.rename(columns={'index': 'cell_id'})

            df_r = pandas2ri.py2rpy(target_pd)
            ro.globalenv['df_r'] = df_r
            # ro.r('result <- frdAllPairsNemenyiTest(df_r$value, df_r$hc, df_r$cell_id)')
            ro.r('result <- frdAllPairsNemenyiTest(df_r$value, df_r$hc, df_r$cell_id)')
            res = ro.r('result')  # Rオブジェクトとして取得
            p_values = res.rx2('p.value')
            p_values = np.vstack(
                [np.full((1, p_values.shape[1]), np.nan), p_values])
            p_values = np.hstack(
                [p_values, np.full((p_values.shape[0], 1), np.nan)])
            p_values = np.where(np.isnan(p_values), p_values.T, p_values)
            p_posthock[region, link] = p_values
            stats_values = res.rx2('statistic')
            stats_values = np.vstack(
                [np.full((1, stats_values.shape[1]), np.nan), stats_values])
            stats_values = np.hstack(
                [stats_values, np.full((stats_values.shape[0], 1), np.nan)])
            stats_values = np.where(
                np.isnan(stats_values), stats_values.T, stats_values)
            stats_posthoc[region, link] = stats_values

            # p_posthock[region, link] = sp.posthoc_dscf(
            #     target_pd.melt(), val_col='value', group_col='variable').to_numpy()

    return activation_rate, chi2_friedman, p_friedman, p_posthock, stats_posthoc, regions, link_typename


def plot_2b(fig, x, y, data, fontsize=8,
            color_dict={"maintained": "blue",
                        "initiated": "orange",
                        "terminated": "green",
                        "transient": "red"}):
    activation_rate, chi2_friedman, p_friedman, p_posthock, stats_posthoc, regions, link_typename = data
    width = 25
    height = 20
    margin_x = 6

    colors = [color_dict[x] for x in link_typename]

    hc_names = ['Pre-conditioning', 'Pre-extinction', 'Post-extinction',
                'Post-retention']

    for region_index, region in enumerate(regions):
        ax = subplot_mm([x+region_index*(width+margin_x),
                        y, width, height], fig=fig)
        sig_bar = {}
        y_max = -np.inf * \
            np.ones(activation_rate[region, link_typename[0]].shape[1])
        for n, link in enumerate(link_typename):
            target = activation_rate[region, link]
            x_val = np.arange(target.shape[1])+(n - 1.5)/20
            y_val = target.mean(axis=0)
            yerr = target.std(axis=0)/np.sqrt(target.shape[0]-1)
            ax.errorbar(
                x_val, y_val, yerr=yerr,
                label=link,
                color=colors[n],
                linewidth=1)
            y_max = np.maximum(y_max, (y_val+yerr))

            if p_friedman[region, link] < 0.05:

                indices = np.where(p_posthock[region, link] < 0.05)
                indices = np.array(
                    [(i, j) for i, j in zip(indices[0], indices[1]) if i <= j])
                for i, j in indices:
                    if (p_posthock[region, link][i, j]) < 0.001:
                        sig_txt = '***'
                    elif (p_posthock[region, link][i, j]) < 0.01:
                        sig_txt = '**'
                    elif (p_posthock[region, link][i, j]) < 0.05:
                        sig_txt = '*'
                    val = [sig_txt, link]
                    if (i, j) in sig_bar.keys():
                        sig_bar[(i, j)].append(val)
                    else:
                        sig_bar[(i, j)] = [val]
        # sig_pos = align_significance_bars(list(sig_bar.keys()))
        sig_pair = list(sig_bar.keys())
        x_pos = np.arange(activation_rate[region, link].shape[1])
        y_pos = y_max*np.ones_like(x_pos)

        ax.set_ylim(0, 32)
        ylim = ax.get_ylim()
        text_y_coord = add_significance_auto(
            ax,
            sig_pair,
            x_pos,
            y_pos,
            significance_text="",
            gap=(ylim[1]-ylim[0])*0.05,
            y_offset=(ylim[1]-ylim[0])*0.075,
            small_gap=(ylim[1]-ylim[0])*0.02,
            fontsize=fontsize*1.5,
            min_vertical_length=(ylim[1]-ylim[0])*0.1,
            base_x_offset=0.05,
            text_offset=(ylim[1]-ylim[0])*(-0.12)
        )
        if sig_bar:
            for sig_x, sig_y in zip(sig_pair, text_y_coord):
                n_sig = len(sig_bar[sig_x])
                x_gap = abs(sig_x[1]-sig_x[0])/(n_sig+1)

                conflict_candidates = [pair for pair, y in zip(
                    sig_pair, text_y_coord) if y > sig_y]
                x_center = (sig_x[1]+sig_x[0])/2
                for conflict_candidate in conflict_candidates:
                    if any([x == x_center for x in conflict_candidate]):
                        conflict = True
                        break
                else:
                    conflict = False

                if len(sig_bar[sig_x]) == 2:
                    jitter = 0.085
                elif conflict:
                    jitter = -0.02
                else:
                    jitter = 0
                for idx, (txt, category) in enumerate(sig_bar[sig_x]):
                    ax.text(
                        min(sig_x)+(idx+1)*x_gap + jitter * (idx - 0.5)*2,
                        sig_y-0.0175*(ylim[1]-ylim[0]),
                        txt,
                        ha='center', va='center',
                        fontsize=fontsize*1.5,
                        color=color_dict[category])

        if region_index == 0:
            ax.set_ylabel("Activation rate\nduring NREM (1/min)")
        else:
            ax.set_ylabel("")
        ax.set_title(region, size=fontsize)
        ax.set_xticks(range(target.shape[1]))
        ax.set_xticklabels(hc_names, rotation=-30, ha='left')
        box_off()

    ax.legend(
        [x.title() for x in link_typename],
        bbox_to_anchor=(1.05, 1),
        loc='upper left',
        borderaxespad=0,
        fontsize=fontsize,
        frameon=False,
        handlelength=1.25,
        handleheight=0.5,
        labelspacing=0.2,
        handletextpad=0.2
    )


def stats_2b(ws, data):
    activation_rate, chi2_friedman, p_friedman, p_posthock, stats_posthoc, regions, link_typename = data
    excel_fonts = excel_font_preset()
    hc_names = ['Pre-conditioning', 'Pre-extinction', 'Post-extinction',
                'Post-retention']
    ws.append(["Region", "Ensemble category", "Session pair", "Statistical Test", "Statistical value",
              "P-value"])

    for cell in ws[1]:
        cell.font = excel_fonts["heading"]

    for region_index, region in enumerate(regions):
        for n, link in enumerate(link_typename):
            target = activation_rate[region, link]
            if p_friedman[region, link] < 0.001:
                p_str = f"{p_friedman[region, link]:.3e}"
            else:
                p_str = f"{p_friedman[region, link]:.3f}"
            ws.append([
                region, link, "NA", f"Friedman test (n = {target.shape[0]})",
                f"χ²₍{to_subscript(target.shape[1]-1)}₎ = {chi2_friedman[region, link]:.2f}",
                p_str])

            if p_friedman[region, link] < 0.05:
                for i, j in itertools.combinations(range(target.shape[1]), 2):
                    if p_posthock[region, link][i, j] < 0.001:
                        p_str = f"{p_posthock[region, link][i, j]:.3e}"
                    else:
                        p_str = f"{p_posthock[region, link][i, j]:.3f}"
                    # ws.append([
                    #     region, link, f"{hc_names[i]} - {hc_names[j]}", f"Post-hoc Steel-Dwass test (n = {target.shape[0]})",
                    #     f"z = {z_posthoc[region, link][i, j]:.2f}", p_str])
                    ws.append([
                        region, link, f"{hc_names[i]} - {hc_names[j]}", f"Post-hoc Nemenyi test (n = {target.shape[0]})",
                        f"q = {stats_posthoc[region, link][i, j]:.2f}", p_str])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = excel_fonts["default"]
    excel_auto_expand(ws)


data_2a = get_data_2a()
data_2b = get_data_2b()

# % %
fontsize = 7
panel_fontsize = 9
plt.rcParams['font.size'] = fontsize
fig_2 = fig_mm([174, 45])
# draw_fig_border()

colors_2a = color_preset('fate_category')

colors_2b = color_preset('fate_category')

x = 1
y = 5

plot_2a(fig_2, x, y, data_2a, fontsize=fontsize, colors=colors_2a)
plot_2a_legend(fig_2, x, y+21, fontsize=fontsize, colors=colors_2a)
text_mm((x, y-2), 'A', fontsize=panel_fontsize, weight='bold', fig=fig_2)


x = 68
y = 7
plot_2b(fig_2, x, y, data_2b, fontsize=fontsize, color_dict=colors_2b)
text_mm((x-13, y-4), 'B', fontsize=panel_fontsize, weight='bold', fig=fig_2)

plt.show()
fig_dir = Path('~/Dropbox/coact_stability_paper/figures').expanduser()
stats_dir = Path('~/Dropbox/coact_stability_paper/stats').expanduser()

if not fig_dir.exists():
    fig_dir.mkdir(parents=True)
fig_2.savefig(fig_dir / f'fig_2_{date.today().strftime("%Y_%m%d")}.pdf')

wb = Workbook()
ws = wb.active
ws.title = "Figure 2A"
stats_2a(ws, data_2a)

ws = wb.create_sheet("Figure 2B")
wb.active = ws
stats_2b(ws, data_2b)

if not stats_dir.exists():
    stats_dir.mkdir(parents=True)
wb.save(stats_dir / f'fig_2_{date.today().strftime("%Y_%m%d")}.xlsx')


# %%

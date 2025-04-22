# %%
from openpyxl import Workbook, load_workbook
from copy import copy
import colorsys
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
# %%


def color_preset(name):
    presets = {
        "fate_category": {
            'maintained': '#D05364',  # D0122B',
            'initiated': '#F5A623',
            'terminated': '#6Ab0F2',
            'transient':  '#3ca33c'  # 2c902c}
        },
        "preserved": {
            'preserved': '#D35F9E',
            'attenuated': '#77C88A'
        },
        "region_pairs": {
            frozenset(('BLA', 'PL5')): '#17BECF',
            frozenset(('PL5', 'vCA1')): '#BCBD22',
            frozenset(('vCA1', 'BLA')): '#FF7F0E'
        },
        "regions": {
            "BLA": "#2bbe27",
            "PL5": "#fa6531",
            "vCA1": "#314dfa"
        },
        "sessions": {
            "Baseline": "#777777",
            "Cue retrieval": "#FF7474",
            "Retention": "#7474FF"
        },
        "homecages": {
            "HC2": "#FF7474",
            "HC3": "#7474FF"
        }
    }

    return presets[name]


def lighten_color(hex_color, brightness=0.9):
    # convert HEX to HLS
    hex_color = hex_color.lstrip("#")
    r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    h, l, s = colorsys.rgb_to_hls(r/255.0, g/255.0, b/255.0)

    l = brightness

    # return HEX code
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return '#{:02X}{:02X}{:02X}'.format(int(r * 255), int(g * 255), int(b * 255))

# %%


def to_subscript(n):
    subscript_map = str.maketrans("0123456789+-=()", "₀₁₂₃₄₅₆₇₈₉₊₋₌₍₎")
    return str(n).translate(subscript_map)


def to_superscript(n):
    superscript_map = str.maketrans("0123456789+-=()", "⁰¹²³⁴⁵⁶⁷⁸⁹⁺⁻⁼⁽⁾")
    return str(n).translate(superscript_map)


def excel_font_preset():
    font_heading = Font(bold=True, name='Times New Roman')
    font_default = Font(bold=False, name='Times New Roman')
    return {"heading": font_heading, "default": font_default}


def excel_auto_expand(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(
            [len(str(cell.value)) for cell in col])+2


def merge_xmlx_sheets(input_files, output_file):

    def copy_sheet_with_styles(source_ws, target_ws):

        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(
                    row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:  # スタイルがある場合のみコピー
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = copy(cell.number_format)
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)

        for col_letter, col_dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter].width = col_dim.width

        for row_idx, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_idx].height = row_dim.height

    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    for file in input_files:
        input_wb = load_workbook(file)
        for sheet_name in input_wb.sheetnames:
            source_ws = input_wb[sheet_name]
            target_ws = output_wb.create_sheet(title=f"{sheet_name}")
            copy_sheet_with_styles(source_ws, target_ws)

    output_wb.save(output_file)
    print(f"Saved to {output_file}")

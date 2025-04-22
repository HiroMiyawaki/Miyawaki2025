# %%
import matplotlib.pyplot as plt
import matplotlib as mpl
import joblib
import numpy as np
from pathlib import Path
from my_tools import data_location
from my_tools import get_fr
from my_tools.pyplot_helper import *
import matplotlib.patches as patches
from datetime import datetime as dt
import pandas as pd
from scipy.signal import find_peaks
from scipy import stats
from matplotlib.colors import LogNorm
import matplotlib.cm as cm
import matplotlib.patches as patches
# LogNormをimport
import math
from my_tools import color_preset
import matplotlib.patches as patches
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font

# %%


def get_data_1c():
    # parameters
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()

    rat = rats[12]
    bin_size = 60.0

    # load data
    _, sleep_state, *_ = \
        joblib.load(data_dir / rat / (rat + "-sleep_states.joblib"))
    _, spikes, cluster_info, *_ \
        = joblib.load(data_dir / rat / (rat + "-ok_spikes.joblib"))
    _, basic_info, session_info, *_ \
        = joblib.load(data_dir / rat / (rat + "-basic_info.joblib"))

    # summarize data
    fr, tbin, c_id = get_fr.in_tbin(
        spikes, basic_info['t_range'], bin_size=bin_size, smooth_sigma=bin_size)
    tbin_edge = np.hstack((tbin - bin_size/2, tbin[-1]+bin_size/2))

    return fr, tbin_edge, cluster_info, basic_info, sleep_state, session_info
# %%


def plot_1c(fig, x, y, data, fontsize=8):
    fr, tbin_edge, cluster_info, basic_info, sleep_state, session_info = data

    width = 149
    h_per_cell = 0.4
    y_gap = 2
    y_stack = y
    gap_for_colorbar = 1
    height_hypnogram = 7
    t_range = [t/3600 for t in basic_info['t_range']]
    regions = ['BLA', 'PL5', 'vCA1']
    for region in regions:
        target = fr[:, cluster_info.query('region==@region').index]
        target = target[:, np.argsort(target.mean(axis=0))]
        target[target == 0] = 1e-2

        c_edge = np.arange(target.shape[1]+1)-0.5
        height = target.shape[1] * h_per_cell
        ax = subplot_mm((x, y_stack, width, height), fig=fig)
        cm = ax.pcolormesh(tbin_edge/3600, c_edge, target.T, norm=LogNorm(
            vmin=1e-1, vmax=5e1), cmap='viridis')
        y_stack = y_stack+height+y_gap
        ax.set_xlim(t_range)
        ax.set_xticklabels([])
        ax.set_yticks([])
        ax.set_ylabel(region)
        box_off()

    ax = subplot_mm((x+width+gap_for_colorbar, y,
                    1.5, y_stack-y_gap-y), fig=fig)
    clim = cm.get_clim()
    c_norm = LogNorm(vmin=clim[0], vmax=clim[1])
    cbar = plt.colorbar(plt.cm.ScalarMappable(c_norm, cm.get_cmap()),
                        cax=ax, orientation="vertical", label='Firing rate (Hz)')
    cbar.set_label
    ax.set_ylabel('Firing rate (Hz)', rotation=-90, va='top')

    ax = subplot_mm((x, y_stack, width, height_hypnogram), fig=fig)
    pos = {'nrem': -0.5, 'rem': 0.5, 'wake': 1.5}
    for s in sleep_state.itertuples():
        rec = patches.Rectangle((s.start_t/3600, pos[s.state]), (s.end_t - s.start_t)/3600,
                                1, facecolor="#000000", edgecolor="#000000", linewidth=0.25)
        ax.add_patch(rec)

    ax.set_xlim(t_range)
    ax.set_ylim([-0.5, 2.5])
    ax.set_xlabel("Time (h)")
    ax.set_yticks(range(3))
    ax.set_yticklabels(['NREM', 'REM', 'WAKE'])
    box_off()

    ax = subplot_mm([x, y, width, y_stack+height_hypnogram-y])
    ax.set_xlim(t_range)
    ax.set_ylim([0, 1])

    target = session_info[session_info.is_behavior_session]
    target = target.query('name != "baseline" & name != "context"')
    target.loc[target['name'] == "cue_and_extinction", 'name'] = 'extinction'
    target.loc[target['name'] ==
               "retention_of_extinction", 'name'] = 'retention'

    for s in target.itertuples():
        rec = patches.Rectangle((s.start_t/3600, 0), (s.end_t - s.start_t)/3600,
                                1, facecolor='none', edgecolor="#ff0066",
                                linewidth=2, clip_on=False)
        ax.add_patch(rec)
        ax.text(s.start_t/3600, 1.01, s.name.replace(
            '_', ' ').capitalize(), color="#ff0066")

    ax.axis('off')
# %%


def get_data_1e():
    # parameters
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()
    rat = rats[12]

    # load data
    _, cos_sim, cos_p, cos_chance, params = joblib.load(
        data_dir / rat / 'analyses' / (rat + '-ic_cossim.joblib'))

    _, ica_weight, *_ = joblib.load(data_dir / rat /
                                    'analyses' / (rat + '-ica_weight.joblib'))

    # summarize data
    template_y = 5
    template_x1 = 1
    template_x2 = 6
    region = 'PL5'
    cim_map1 = cos_sim[region][template_y][template_x1]
    cim_map2 = cos_sim[region][template_y][template_x2]
    p_map1 = cos_p[region][template_y][template_x1]
    p_map2 = cos_p[region][template_y][template_x2]

    w_map_y = np.vstack(ica_weight[template_y].query('region==@region').weight)
    w_map_x1 = np.vstack(
        ica_weight[template_x1].query('region==@region').weight)
    w_map_x2 = np.vstack(
        ica_weight[template_x2].query('region==@region').weight)

    y_idx_orig = np.arange(cim_map1.shape[0])
    x1_idx_orig = np.arange(cim_map1.shape[1])
    x2_idx_orig = np.arange(cim_map2.shape[1])
    for n in range(min(cim_map1.shape)-1):
        am = np.unravel_index(
            np.argmax(cim_map1[n:, n:]), cim_map1[n:, n:].shape)

        y_idx = list(range(cim_map1.shape[0]))
        x1_idx = list(range(cim_map1.shape[1]))
        y_idx[am[0]+n] = n
        x1_idx[am[1]+n] = n
        y_idx[n] = am[0]+n
        x1_idx[n] = am[1]+n

        cim_map1 = cim_map1[y_idx, :]
        cim_map1 = cim_map1[:, x1_idx]
        cim_map2 = cim_map2[y_idx, :]
        p_map1 = p_map1[y_idx, :]
        p_map1 = p_map1[:, x1_idx]
        p_map2 = p_map2[y_idx, :]
        w_map_y = w_map_y[y_idx, :]
        w_map_x1 = w_map_x1[x1_idx, :]

        y_idx_orig = y_idx_orig[y_idx]
        x1_idx_orig = x1_idx_orig[x1_idx]

        if n < cim_map2.shape[1]:
            am = np.argmax(cim_map2[n, n:])
            x2_idx = list(range(cim_map2.shape[1]))
            x2_idx[am+n] = n
            x2_idx[n] = am+n
            cim_map2 = cim_map2[:, x2_idx]
            p_map2 = p_map2[:, x2_idx]
            w_map_x2 = w_map_x2[x2_idx, :]
            x2_idx_orig = x2_idx_orig[x2_idx]

    sorted_cell_idx = np.empty((0), dtype=int)
    am = np.abs(w_map_y).argmax(axis=0)
    for n in range(w_map_y.shape[0]):
        target = np.array(np.where(am == n))[0]
        target = target[np.argsort(w_map_y[n, target])[::-1]]
        sorted_cell_idx = np.hstack((sorted_cell_idx, target))

    w_map_y = w_map_y[:, sorted_cell_idx]
    w_map_x1 = w_map_x1[:, sorted_cell_idx]
    w_map_x2 = w_map_x2[:, sorted_cell_idx]

    w_map_x2[0, :] = - w_map_x2[0, :]

    return \
        cim_map1, cim_map2, p_map1, p_map2, \
        w_map_y, w_map_x1, w_map_x2, \
        cos_chance, region, rat, template_y, \
        y_idx_orig, sorted_cell_idx


def plot_1e(fig, x, y, data, fontsize=8,
            targets=None, ensemble_colors=None):
    cell_height = 2
    weight_total_size = 15
    gap = 1
    width = 0.5

    cim_map1, cim_map2, \
        p_map1, p_map2, \
        w_map_y, w_map_x1, \
        w_map_x2, cos_chance, region, *_ = data

    w_range = [-0.8, 0.8]
    sim_range = [0, 0.7]

    left_pos = x
    top_pos = y + weight_total_size+gap
    width = weight_total_size
    height = cim_map1.shape[0]*cell_height
    ax = subplot_mm([left_pos, top_pos, width, height], fig=fig)

    ax.pcolormesh(w_map_y, cmap=plt.cm.bwr, vmax=w_range[1], vmin=w_range[0])
    ax.invert_yaxis()
    ax.set_yticks([])
    ax.set_ylabel('Extinction-\nensembles', labelpad=0, fontsize=fontsize)
    ax.set_xticks(range(10, w_map_x1.shape[1], 10))
    ax.set_xlabel('Cell ID', fontsize=fontsize)
    ax.tick_params(length=2)

    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    if targets is not None:
        print(targets)
        for n in range(len(targets)):
            arrow_color = ensemble_colors[n]  # 矢印の色
            arrow = patches.FancyArrow(
                x=-4.5,
                y=targets[n] + 0.5,  # セルの中央に矢印を配置
                dx=3,
                dy=0,  # 水平方向に矢印を引く
                width=0.05,  # 矢印の幅
                head_width=0.3,  # 矢印の頭の幅
                head_length=1.5,  # 矢印の頭の長さ
                color=arrow_color,
                length_includes_head=True,
                clip_on=False
            )
            ax.add_patch(arrow)
            # ax.annotate(
            #     "",  # テキストなし
            #     xy=(-2, targets[n] + 0.5),  # 矢頭の位置
            #     xytext=(-10, targets[n] + 0.5),  # 基準位置
            #     arrowprops=dict(
            #         arrowstyle="->",  # 矢頭のみを描画
            #         color=ensemble_colors[n],  # 矢頭の色
            #         lw=1.0  # 線の太さ
            #         ),
            #     clip_on=False)  # 枠外に描画を許可

    left_pos = x+weight_total_size+gap
    top_pos = y
    w = cim_map1.shape[1]*cell_height
    height = weight_total_size
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    ax.pcolormesh(
        w_map_x1.T, cmap=plt.cm.bwr,
        vmax=w_range[1], vmin=w_range[0])
    ax.invert_yaxis()
    ax.set_xticks([])
    ax.xaxis.set_label_position('top')
    ax.set_xlabel('Conditioning-\nensembles', fontsize=fontsize)
    ax.set_yticks(range(10, w_map_x1.shape[1], 10))
    ax.set_ylabel('Cell ID', fontsize=fontsize)
    ax.tick_params(length=2)
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    # change_ax_width(ax, width)

    left_pos = x+weight_total_size+gap
    top_pos = y+weight_total_size+gap
    height = cim_map1.shape[0]*cell_height
    w = cim_map1.shape[1]*cell_height
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    ax.pcolormesh(
        cim_map1, cmap=plt.cm.inferno,
        vmax=sim_range[1], vmin=sim_range[0])
    ax.invert_yaxis()
    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    # change_ax_width(ax, width)

    for sig_pos in np.array(np.where((p_map1 < 0.01) & (cim_map1 > np.median(cos_chance[region])))).T:
        ax.text(*(sig_pos[::-1]+[0.55, 0.8]), '**',
                ha='center', va='center', fontsize=fontsize)

    left_pos = x+weight_total_size+gap+cim_map1.shape[1]*cell_height+gap
    top_pos = y
    w = cim_map2.shape[1]*cell_height
    height = weight_total_size
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    ax.pcolormesh(
        w_map_x2.T, cmap=plt.cm.bwr,
        vmax=w_range[1], vmin=w_range[0])
    ax.invert_yaxis()
    ax.set_yticks([])
    ax.set_xticks([])
    ax.xaxis.set_label_position('top')
    ax.set_xlabel('Retention-\nensembles', fontsize=fontsize)
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    left_pos = x+weight_total_size+gap+cim_map1.shape[1]*cell_height+gap
    top_pos = y+weight_total_size+gap
    height = cim_map2.shape[0]*cell_height
    w = cim_map2.shape[1]*cell_height
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    ax.pcolormesh(
        cim_map2, cmap=plt.cm.inferno,
        vmax=sim_range[1], vmin=sim_range[0])
    ax.invert_yaxis()
    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    for sig_pos in np.array(np.where((p_map2 < 0.01) & (cim_map2 > np.median(cos_chance[region])))).T:
        ax.text(*(sig_pos[::-1]+[0.55, 0.8]), '**',
                ha='center', va='center', fontsize=fontsize)

    left_pos = x+weight_total_size + \
        (cim_map1.shape[1]+cim_map2.shape[1])*cell_height+gap*3
    top_pos = y+weight_total_size+gap
    height = cim_map2.shape[0]*cell_height
    w = 1.5
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    val = np.linspace(*sim_range, 256)
    step = np.diff(val).mean()
    edge = np.hstack((val, val[-1]+step/2))-step/2
    [xx, yy] = np.meshgrid([0, 1], edge)
    val = val.reshape((-1, 1))
    ax.pcolormesh(xx, yy, val, cmap=plt.cm.inferno)
    ax.yaxis.set_label_position('right')
    ax.yaxis.set_ticks_position('right')
    ax.set_xticks([])
    ax.set_ylabel('Absolute\ncosine simirarity', rotation=-
                  90, va='bottom', fontsize=fontsize)
    ax.tick_params(length=2)
    ax.collections[0].set_edgecolor('face')
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

    top_pos = y
    height = weight_total_size
    ax = subplot_mm([left_pos, top_pos, w, height], fig=fig)
    val = np.linspace(*w_range, 256)
    step = np.diff(val).mean()
    edge = np.hstack((val, val[-1]+step/2))-step/2
    [xx, yy] = np.meshgrid([0, 1], edge)
    val = val.reshape((-1, 1))
    ax.pcolormesh(xx, yy, val, cmap=plt.cm.bwr)
    ax.yaxis.set_label_position('right')
    ax.yaxis.set_ticks_position('right')
    ax.set_xticks([])
    ax.set_ylabel('Weight', rotation=-90, va='bottom', fontsize=fontsize)
    ax.tick_params(length=2)
    ax.collections[0].set_edgecolor('face')
    for spine in ax.spines.values():
        lw = spine.get_linewidth()   # 線幅を取得
        spine.set_position(('outward', lw/2))

# %%


def get_data_1d(data):
    # get parameters
    *_, region, rat, template_y, \
        y_idx_orig, sorted_cell_idx = data

    data_dir = data_location.get_rootdir()

    # load data
    _, ica_weight, *_ = joblib.load(data_dir / rat /
                                    'analyses' / (rat + '-ica_weight.joblib'))

    _, spikes, cluster_info, *_ \
        = joblib.load(data_dir / rat / (rat + "-ok_spikes.joblib"))

    _, react_strength, react_param = joblib.load(data_dir / rat / 'analyses' /
                                                 (rat + "-ica_react_strength-cue_and_extinction-ext.joblib"))

    react_time = (
        np.arange(react_strength.shape[0])+0.5)*react_param["tbin_size"]

    # sort cells and ensembles
    rev_sorted_cell_idx = sorted_cell_idx[::-1]
    weight_vectors = ica_weight[template_y]
    ensemble_id = weight_vectors.index[weight_vectors["region"] == region].to_numpy()[
        y_idx_orig]

    weight = weight_vectors.loc[ensemble_id, "weight"].values
    weight = [x[rev_sorted_cell_idx] for x in weight]
    cell_ids_orig = np.array((weight_vectors.loc[ensemble_id, "cell_id"].values)[0])[
        rev_sorted_cell_idx]

    # get subsets of spike data
    subset = spikes[spikes["cluster"].isin(cell_ids_orig)]
    spk = subset["spiketime"].values
    clu = subset["cluster"].values

    mapping = {cell_id: idx for idx, cell_id in enumerate(cell_ids_orig)}
    clu = np.array([mapping[val] for val in clu])

    return \
        react_strength, react_time, ensemble_id, \
        weight, clu, spk, cell_ids_orig


def plot_1d(fig, x, y, data, targets=[0, 8], fontsize=8,
            ensemble_colors=np.array(["#4CAF50", "#F44336", "#555555"]),
            t_start=36393, dur=3):
    # set parameters
    raster_width = 70
    height = 35
    weight_width = 10
    ensemble_height = 5

    # 36231 36378 36393 36897 37089 37113 37610 37792 38052 38316 38349 38358 38493 39141
    # t_start = 36378
    t_end = t_start + dur

    # ensemble_colors = ['r', 'b', 'g', 'y']
    # ensemble_colors = np.array([[0, 0, 0], [1, 0, 0], [0, 0, 1], [1, 0, 1]])

    targets = np.array(targets)

    react_strength, react_time, ensemble_id, \
        weight, clu, spk, cell_ids_orig = data

    col_idx = 2*np.zeros_like(cell_ids_orig)
    z_scored_weight = [(x - x.mean())/x.std() for x in weight]
    contribute = np.zeros_like(cell_ids_orig)
    for n, target in enumerate(targets):
        contribute[z_scored_weight[target] > 2] += 2**n
    col_idx[contribute == 0] = 2
    col_idx[contribute == 1] = 0
    col_idx[contribute == 2] = 1
    target_react = react_strength[:, ensemble_id[targets]]

    ax = subplot_mm([x, y+ensemble_height, raster_width, height], fig=fig)

    ax.set_xlim([t_start, t_end])
    t_start = t_start-0.04
    t_end = t_end+0.04

    target_react = target_react[(react_time > t_start)
                                & (react_time < t_end), :]
    target_react_time = react_time[(
        react_time > t_start) & (react_time < t_end)]
    peak_t = []
    rect_bottom = []
    rect_top = []
    for n in range(len(targets)):
        contributing_cells = np.where([(x >> n) & 1 for x in contribute])[0]
        contributing_idx_min = contributing_cells.min()
        contributing_idx_max = contributing_cells.max()
        peaks = find_peaks(target_react[:, n], height=5)
        for peak in peaks[0]:
            peak_t.append(target_react_time[peak])
            rect_bottom.append(contributing_idx_min-0.5)
            rect_top.append(contributing_idx_max+0.5)

    y_val = clu[(spk > t_start) & (spk < t_end)]
    x_val = spk[(spk > t_start) & (spk < t_end)]

    ax.scatter(x_val, y_val, c=ensemble_colors[col_idx[y_val]], s=1, zorder=5)

    for n in range(len(peak_t)):
        ax.add_patch(patches.Rectangle(
            (peak_t[n]-0.02, rect_bottom[n]),
            0.04, rect_top[n]-rect_bottom[n],
            fill=False,
            edgecolor='black',
            lw=0.5,
            zorder=1))

    y_range = list(ax.get_ylim())
    y_range[1] = len(cell_ids_orig)+0.5
    ax.set_ylim(y_range)
    x_range = ax.get_xlim()
    ax.set_axis_off()
    # ax.set_title(f"{t_start+0.04:.2f} - {t_end-0.04:.2f}")

    ax = subplot_mm([x+raster_width, y+ensemble_height,
                    weight_width, height], fig=fig)
    gap = 0.1
    ax.plot([0, 0], y_range, 'k-', lw=1)
    for n, target in enumerate(targets):
        sig_color = ensemble_colors[n]
        ns_color = color_preset.lighten_color(sig_color, 0.8)
        for m in range(len(weight[target])):
            if (contribute[m] >> n) & 1 == 1:
                col = sig_color
            else:
                col = ns_color

            ax.plot([0, weight[target][m]],
                    [m+2*(n-0.5)*gap, m+2*(n-0.5)*gap],
                    color=col, lw=0.75)
            ax.scatter(weight[target][m], m+2*(n-0.5)*gap, 2, c=col)
        # for m in range(len(weight[n])):
        #     ax.plot([0, weight[n][m]],
        #             [m+2*(n-0.5)*gap, m+2*(n-0.5)*gap],
        #             color=col, lw=0.75)
    ax.set_ylim(y_range)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.set_yticks([])
    ax.set_xlabel('Weight')

    ax = subplot_mm([x, y,
                    raster_width, ensemble_height], fig=fig)

    for n in range(len(targets)):
        ax.plot(target_react_time,
                target_react[:, n]-n*5, '-', color=ensemble_colors[n], lw=0.75)

    y_range = ax.get_ylim()
    bar_left = np.array(x_range)@np.array([0.9, 0.1])
    bar_bottom = np.array(y_range)@np.array([0.55, 0.45])
    ax.plot(bar_left+np.array([0, dur/10]), [bar_bottom, bar_bottom], 'k-')
    ax.text(bar_left+dur/20, bar_bottom+3,
            f'{dur*100:.0f} ms', ha='center', va='bottom')

    bar_left = np.array(x_range)@np.array([0.93, 0.07])
    bar_bottom = np.array(y_range)@np.array([0.5, 0.5])
    ax.plot([bar_left, bar_left], bar_bottom+np.array([0, 10]), 'k-')
    ax.text(bar_left-0.04, bar_bottom+5, '10 z', ha='right', va='center')
    ax.set_xlim(x_range)
    ax.set_axis_off()

# %%

# %%


def get_data_1b():
    data_dir = data_location.get_rootdir()
    rats = data_location.get_rats()
    freeze_duration = np.zeros((len(rats), 5), dtype=object)
    total_duration = np.zeros((len(rats), 5), dtype=object)

    for rat_idx, rat in enumerate(rats):

        _, freezes_df, _ = joblib.load(data_dir / rat / f'{rat}-freeze.joblib')
        _, cues_df, _ = joblib.load(data_dir / rat / f'{rat}-cues.joblib')
        _, sleep_df, *_ = joblib.load(data_dir /
                                      rat / f'{rat}-sleep_states.joblib')

        _, basic_info, session_df, *_ \
            = joblib.load(data_dir / rat / f'{rat}-basic_info.joblib')

        for ses_idx in range(5):
            if ses_idx == 2:
                continue

            session_t = session_df[
                session_df["is_behavior_session"]][["start_t", "end_t"]].to_numpy()

            target_time = session_t[ses_idx, :]

            sleep_subset = sleep_df[(sleep_df["start_t"] < target_time[1]) & (
                sleep_df["end_t"] > target_time[0])]
            exclude_t = sleep_subset[
                sleep_subset["state"] != "wake"][["start_t", "end_t"]].to_numpy()

            cue_subset = cues_df[cues_df["session_idx"] == ses_idx]

            cue_t = pd.merge(
                cue_subset.loc[
                    cue_subset.groupby("train_idx")["pip_idx"].idxmin(),
                    ["train_idx", "start_t"]],
                cue_subset.loc[
                    cue_subset.groupby("train_idx")["pip_idx"].idxmax(),
                    ["train_idx", "end_t"]],
                on="train_idx")[["start_t", "end_t"]].to_numpy()

            freeze_t = freezes_df[
                (freezes_df["end_t"] > target_time[0]) &
                (freezes_df["start_t"] < target_time[1])].to_numpy()

            t_edge = np.sort(cue_t.reshape(-1))
            t_edge = np.append(t_edge, target_time)

            frz_dur = np.zeros(len(t_edge)-1)
            bin_dur = np.zeros(len(t_edge)-1)

            for n in range(len(t_edge)-1):
                ex_sub = exclude_t[(exclude_t[:, 0] < t_edge[n+1])
                                   & (exclude_t[:, 1] > t_edge[n]), :]

                if ex_sub.shape[0] > 0:
                    if ex_sub[0, 0] < t_edge[n]:
                        ex_sub[0, 0] = t_edge[n]
                    if ex_sub[-1, 1] > t_edge[n+1]:
                        ex_sub[-1, 1] = t_edge[n+1]

                frz_sub = freeze_t[
                    (freeze_t[:, 0] < t_edge[n+1]) & (freeze_t[:, 1] > t_edge[n]), :]
                if frz_sub.shape[0] > 0:
                    if frz_sub[0, 0] < t_edge[n]:
                        frz_sub[0, 0] = t_edge[n]
                    if frz_sub[-1, 1] > t_edge[n+1]:
                        frz_sub[-1, 1] = t_edge[n+1]

                dur = 0
                for frz in frz_sub:
                    ex = ex_sub[(ex_sub[:, 0] < frz[1]) &
                                (ex_sub[:, 1] > frz[0]), :]
                    if ex.shape[0] > 0:
                        if ex[0, 0] < frz[0]:
                            ex[0, 0] = frz[0]
                        if ex[-1, 1] > frz[1]:
                            ex[-1, 1] = frz[1]
                    dur += frz[1] - frz[0] - np.sum(ex[:, 1] - ex[:, 0])

                frz_dur[n] = dur
                bin_dur[n] = t_edge[n+1] - t_edge[n]

                freeze_duration[rat_idx, ses_idx] = frz_dur
                total_duration[rat_idx, ses_idx] = bin_dur

    n_cue_def = 3
    frz = np.zeros((len(rats), 5, n_cue_def*2))
    dur = np.zeros((len(rats), 5, n_cue_def*2))
    pool_rate = np.zeros((5, 4), dtype=object)
    for target_idx in range(5):
        for n in range(4):
            pool_rate[target_idx, n] = []

        n_cue = n_cue_def
        if target_idx == 0:  # baseline
            ses_idx = 0
            cue_offset = 0
            if n_cue > 4:
                n_cue = 4
        elif target_idx == 1:  # conditioning
            ses_idx = 1
            cue_offset = 0
            if n_cue > 12:
                n_cue = 12
        elif target_idx == 2:  # cue retention
            ses_idx = 3
            cue_offset = 0
            if n_cue > 8:
                n_cue = 8
        elif target_idx == 3:  # extinction
            ses_idx = 3
            cue_offset = 2 * 8 + 1
        elif target_idx == 4:  # retention of extinction
            ses_idx = 4
            cue_offset = 0
            if n_cue > 8:
                n_cue = 8

        for rat_idx in range(len(rats)):
            frz[rat_idx, target_idx, :] = freeze_duration[rat_idx,
                                                          ses_idx][cue_offset:cue_offset+n_cue*2]
            dur[rat_idx, target_idx, :] = total_duration[rat_idx,
                                                         ses_idx][cue_offset:cue_offset+n_cue*2]

    f_rate = frz[:, :, :]/dur[:, :, :]*100
    avg = np.nanmean(f_rate, axis=0)
    ste = np.nanstd(f_rate, axis=0)/(len(rats)**0.5)

    r = frz[:, :, :].sum(axis=2)/dur[:, :, :].sum(axis=2)*100
    t_w, p_w = stats.wilcoxon(r[:, 2], r[:, 4])

    freeze_mean = avg[[0, 2, 4], :]
    freeze_ste = ste[[0, 2, 4], :]
    freeze_rate = r[:, [2, 4]]
    freeze_p = p_w
    freeze_t = t_w

    return freeze_mean, freeze_ste, freeze_rate, freeze_p, freeze_t

# %%


def plot_1b(fig, x, y, data, fontsize=8,
            color={
                "Baseline": "#777777",
                "Cue retrieval": "#FF7474",
                "Retention": "#7474FF"}):
    freeze_mean, freeze_ste, freeze_rate, freeze_p, freeze_t = data
    session = ["Baseline", "Cue retrieval", "Retention"]
    trisec_width = 25
    mean_width = 7
    legend_width = 15
    x_gap = 26
    height = 20
    y_range = [0, 110]
    ax = subplot_mm([x, y, trisec_width, height], fig=fig)
    for n in range(3):
        ax.errorbar(np.arange(6)+(n-1)*0.1, freeze_mean[n, :],
                    yerr=freeze_ste[n, :], fmt='-',
                    color=color[session[n]], label=session[n], zorder=3)

    for idx in range(3):
        rect = patches.Rectangle(
            [2*idx-0.5, 0], 1, 110, facecolor="#FFDD99", linestyle="none", zorder=1)
        ax.add_patch(rect)
    ax.set_xticks([0, 2, 4])
    ax.set_xticklabels(["1st", "2nd", "3rd"])
    ax.tick_params(axis="x", length=0)
    ax.set_ylim(y_range)
    ax.set_xlim([-0.5, 6.2])
    ax.set_ylabel("Freeze (%)")
    ax.set_xlabel("Cue presentation")
    box_off(ax)
    ax = subplot_mm([x+trisec_width, y, legend_width, height], fig=fig)
    ax.set_xlim([0, legend_width])
    ax.set_ylim([height, 0])
    ones = np.ones(2)
    ax.plot([1, 4], 1*ones, color=color["Cue retrieval"])
    ax.text(1, 2, "Cue-\nretrieval", ha="left", va="top",
            color=color["Cue retrieval"], fontsize=fontsize)
    ax.plot([1, 4], 9*ones, color=color["Retention"])
    ax.text(1, 10, "Retention", ha="left", va="top",
            color=color["Retention"], fontsize=fontsize)
    ax.plot([1, 4], 17*ones, color=color["Baseline"])
    ax.text(1, 18, "Baseline", ha="left", va="top",
            color=color["Baseline"], fontsize=fontsize)
    ax.axis("off")
    ax = subplot_mm([x+trisec_width+x_gap, y, mean_width, height], fig=fig)
    avg = freeze_rate.mean(axis=0)
    ste = freeze_rate.mean(axis=0)/(freeze_rate.shape[0]**0.5)
    for n in range(len(avg)):
        ax.bar(n, avg[n], color=color[session[n+1]])
        ax.plot([n, n], [avg[n], avg[n]+ste[n]], '-', lw=1.5,
                color=color[session[n+1]])
        # ax.errorbar(n, avg[n], yerr=ste[n], fmt='-')
    ax.plot(range(2), freeze_rate.T, 'o', markersize=2, markerfacecolor='none',
            markeredgecolor="#666666", markeredgewidth=0.5)
    ax.set_ylim(y_range)
    box_off(ax)
    ax.set_xticks(range(2))
    ax.set_xticklabels(session[1:], rotation=-30, ha="left")

    if freeze_p < 0.001:
        p_str = "***"
    elif freeze_p < 0.01:
        p_str = "**"
    elif freeze_p < 0.05:
        p_str = "*"
    else:
        p_str = ""
    ax.plot([0, 0, 1, 1], [102, 107, 107, 102], 'k', lw=0.5)
    ax.text(0.5, 107, p_str, ha='center', va='center', fontsize=fontsize*1.5)
    ax.set_ylabel("Freeze (%)")


# %%
def stats_1b(ws, data):
    freeze_mean, freeze_ste, freeze_rate, freeze_p, freeze_t = data
    excel_fonts = color_preset.excel_font_preset()
    ws.append(["Session", "Statistical Test", "Statistical value",
              "P-value"])

    for cell in ws[1]:
        cell.font = excel_fonts["heading"]

    if freeze_p < 0.001:
        p_str = f"{freeze_p:.3e}"
    else:
        p_str = f"{freeze_p:.3f}"

    ws.append([
        "[Cue retrieval] vs [Retention]", f"Wilcoxon rank sum test (n = {freeze_rate.shape[0]})",
        f"T = {freeze_t:.2f}", p_str])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = excel_fonts["default"]
    color_preset.excel_auto_expand(ws)
# %%


def plot_1a(fig, x, y, width=84, height=18, fontsize=8):
    # subplot_mm はユーザー定義の関数と仮定
    ax = subplot_mm([x, y, width, height], fig=fig)
    ax.set_facecolor('lightgray')

    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.tick_params(left=False, right=False, bottom=False, top=False,
                   labelleft=False, labelbottom=False)


# %%
data_1b = get_data_1b()
data_1c = get_data_1c()
data_1e = get_data_1e()
data_1d = get_data_1d(data_1e)
# %%

fontsize = 7
panel_fontsize = 9
plt.rcParams['font.size'] = fontsize
fig_1 = fig_mm([174, 140])
# draw_fig_border()

x = 10
y = 7
plot_1a(fig_1, x, y)
text_mm((x-6, y-1), 'A', fontsize=panel_fontsize, weight='bold', fig=fig_1)

x = 107
y = 5
session_colors = color_preset.color_preset("sessions")
plot_1b(fig_1, x, y, data_1b, fontsize=fontsize, color=session_colors)
text_mm((x-12, y+1), 'B', fontsize=panel_fontsize, weight='bold', fig=fig_1)

x = 10
y = 38
plot_1c(fig_1, x, y, data_1c)
text_mm((x-6, y), 'C', fontsize=panel_fontsize, weight='bold', fig=fig_1)

targets = [0, 1]
t_start = 38493.3
dur = 2.5
ensemble_colors = np.array(["#4CAF50", "#F44336", "#555555"])

x = 10
y = 91
plot_1d(fig_1, x, y, data_1d,
        targets=targets, t_start=t_start, dur=dur, ensemble_colors=ensemble_colors)
text_mm((x-6, y+2), 'D', fontsize=panel_fontsize, weight='bold', fig=fig_1)

x = 100
y = 95
plot_1e(fig_1, x, y, data_1e,
        targets=targets, ensemble_colors=ensemble_colors, fontsize=fontsize)
text_mm((x-3, y-2), 'E', fontsize=panel_fontsize, weight='bold', fig=fig_1)


plt.show()
fig_dir = Path('~/Dropbox/coact_stability_paper/figures').expanduser()
stats_dir = Path('~/Dropbox/coact_stability_paper/stats').expanduser()

if not fig_dir.exists():
    fig_dir.mkdir(parents=True)
fig_1.savefig(fig_dir / f'fig_1_{date.today().strftime("%Y_%m%d")}.pdf')

wb = Workbook()
ws = wb.active
ws.title = "Figure 1B"
stats_1b(ws, data_1b)


if not stats_dir.exists():
    stats_dir.mkdir(parents=True)
wb.save(stats_dir / f'fig_1_{date.today().strftime("%Y_%m%d")}.xlsx')
